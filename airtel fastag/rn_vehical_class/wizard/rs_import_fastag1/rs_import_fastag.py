# -*- coding: utf-8 -*-
###################################################################################
#
#    Redian Software Pvt. Ltd
#
#    Copyright (C) 2019-TODAY Redian Software(<https://www.rediansoftware.com/>).
#    This program is free software: you can modify
#    it under the terms of the GNU Affero General Public License (AGPL) as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
###################################################################################

from odoo import models, fields, api, _
import base64
import xlrd
import tempfile
import binascii
import io
from datetime import datetime, date
from odoo.exceptions import UserError, ValidationError, RedirectWarning, Warning
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import csv
import subprocess
from pyexcel_ods import get_data
import codecs

class OrderLineProduct(models.TransientModel):
    _name = "order.line.product"

    fastag_product = fields.Binary('Select File')
    # fastag_product = fields.Char('Select File')
    order_line_form = fields.One2many('fastag.line.form', 'product_line_id',
                                      string="All Fastags")
    user_id = fields.Many2one('res.users', string="User",
                              default=lambda self: self.env.user.id)
    picking_id = fields.Many2one('stock.picking', "Picking",
                                 ondelete='cascade')

    def upload_product_line(self):
        if self.fastag_product == False:
            raise UserError(
                _('Please Upload your excel first then imoport fastag'))
        else:
            csv_data = base64.b64decode(self.fastag_product)
            string_data = csv_data.decode('utf-7')
            data_file = io.StringIO(string_data)
            data_file.seek(0)
            file_reader = []
            try:
                csv_reader = csv.reader(data_file, delimiter=',')
                next(csv_reader)
                file_reader.extend(csv_reader)
                data = len(list(file_reader))
                picking_id = self.env['stock.picking'].browse(
                    self._context.get('active_id', False))
                today = date.today()
                lst_date = str(today).split("-")
                today = lst_date[2] + "/" + lst_date[1] + "/" + lst_date[0]
                for row_vals in file_reader:
                    print('file_reader',str(row_vals[0]))
                    print(data)
                    terms = []
                    if picking_id:
                        login = row_vals[5]
                        assigned_id = self.env['res.users'].sudo().search(
                            [('login', '=', login)])
                        if assigned_id:
                            # picking_id.sudo().write({'rs_emp_no': int(row_vals[5]),
                            #                          'rs_destination_id': assigned_id.id,
                            #
                            #                          })
                            query1 = """SELECT * FROM rs_circle_name WHERE
                                                                     name = %(circle)s"""
                            self.env.cr.execute(query1, {
                                                        'circle': str(row_vals[1]),
                                                        })
                            circle_id = self.env.cr.fetchall()
                            print('circle',circle_id[0][0])
                            query = """UPDATE stock_picking
                                                    SET rs_emp_no = %(emp)s,
                                                        rs_destination_id = %(dest)s,
                                                        rs_circle_id = %(circle)s
                                                     WHERE id = %(name)s """
                            self.env.cr.execute(query, {'emp': int(row_vals[5]),
                                                        'dest': assigned_id.id,
                                                        'circle': circle_id[0][0],
                                                        'name': picking_id.id
                                                        })

                        if assigned_id.rs_designation == 'fastag_promoter':
                            records = self.env[
                                'product.template'].sudo().search_count(
                                [('warehouse_emp_no', '=', assigned_id.rs_employee_id)])

                            # fastags_needed = int(records) + int(len(data))
                            fastags_needed = int(records) + data
                            if int(records) >= 100:
                                raise UserError(
                                    _('Promoter %s fastag count is exceeds the limit 100',
                                      assigned_id.name))
                            elif int(fastags_needed) > 100:
                                raise UserError(
                                    _('Promoter %s Already there are %s FASTag you can assign %s !',
                                      assigned_id.name, int(records),
                                      data - 100))
                    product_ids = self.env['product.product'].sudo().search(
                        [('barcode', '=', str(row_vals[0]))])
                    print('================================',str(row_vals[0]))
                    pro = self.env['product.template'].sudo().search(
                        [('barcode', '=', str(row_vals[0]))])

                    product_count = self.env['product.template'].search_count([])
                    print('product counttttttttttttttttttttt',product_count)

                    if str(product_ids.barcode) != str(row_vals[0]):
                        raise UserError(
                            _('%s Barcode is not available in the Products !',
                              str(row_vals[0])))
                    for pro in pro:
                        if str(pro.barcode) == str(row_vals[0]):
                            print(pro.barcode)
                            if pro.circle_name.name != row_vals[1]:
                                print(pro.circle_name.name)
                                raise UserError(
                                    _('The product with %s having defferent circle name %s',
                                      str(row_vals[0]), str(row_vals[1])))

                    sold_count = self.env['product.template'].search_count(
                        [('fastag_sold', '=', 'yes')])
                    sold_pro = product_count - sold_count
                    if data > sold_pro:
                        raise UserError(
                            _('Only %s fastag is available, %s are sold fastag',
                              sold_pro, sold_count))

                    fault_count = self.env['product.template'].search_count(
                        [('rs_faulty_stag', '=', 'yes')])
                    fault_pro = product_count - fault_count
                    if data > fault_pro:
                        raise UserError(
                            _('Only %s fastag is available, %s are faulty fastag',
                              fault_pro, fault_count))

                    unlink_count = self.env['product.template'].search_count(
                        [('unlink_fastag', '=', True)])
                    unlink_pro = product_count - unlink_count
                    if data > unlink_pro:
                        raise UserError(
                            _('Only %s fastag is available, %s are unlink fastag',
                              unlink_pro, unlink_count))

                    if data > 101 and assigned_id.rs_designation == 'fastag_promoter':
                        raise UserError(
                            _('You are not allowed to trasfer more then 100 fastag !'))
                    if data > 1001 and assigned_id.rs_designation == 'fastag_tl':
                        raise UserError(
                            _('You are not allowed to trasfer more then 1000 fastag !'))
                    product_circle_count = self.env['product.template'].sudo().search_count(
                        [('circle_name', '=', row_vals[1])])

                    if data > product_circle_count:
                        raise UserError(
                            _('Not enough number of fastag for this circle'))
                    product_tag_count = self.env[
                        'product.template'].sudo().search_count(
                        [('categ_id', '=', product_ids.categ_id.id)])
                    if data > product_tag_count:
                        raise UserError(
                            _('Not enough number of fastag for this tag class'))
                    for product_id in product_ids:
                        terms.append((0, 0, {'picking_id': picking_id.id,
                                             'product_id': product_id.id,
                                             'fatag_bracode': str(row_vals[0]),
                                             'circle_name': row_vals[1],
                                             'date_of_dispatch': today,
                                             'consignment_no': str(row_vals[3]),
                                             'delivery_partner': row_vals[4],
                                             'emp_mobile_no': row_vals[5],
                                            'category_id': product_id.categ_id.id
                                            #'category_id': int(row_vals[6])
                                             }))
                    self.sudo().update({'order_line_form': terms})
                return {
                    'name': _('Product Selection'),
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'order.line.product',
                    'view_id': self.env.ref(
                        'rn_vehical_class.add_product_in_line_form_view').id,
                    'res_id': self.id,
                    'target': 'new'
                }
            except csv.Error:
                raise UserError(
                    _("Cannot determine the file format for the attached file."))

    def product_line(self):
        # product = self.env['product.template'].search(
        #     [('fastag_sold', '=', 'yes')])
        # for product in product:
        #     query1 = """UPDATE product_template
        #                                            SET fastag_sold = 'no'
        #                                             WHERE id = %(name)s """
        #     self.env.cr.execute(query1, {'name': product.id
        #                                  })
        if self.fastag_product == False:
            raise UserError(
                _('Please Upload your excel first then imoport fastag'))
        else:
            # terms = []
            # file_datas = base64.decodestring(self.fastag_product)
            # workbook = xlrd.open_workbook(file_contents=file_datas)
            #
            # #        Read worksheet
            # sheet = workbook.sheet_by_index(0)
            # #        Prepare list of values
            #
            # data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r
            #         in range(sheet.nrows)]
            # # picking_id = self.env['stock.picking'].browse(
            # #     self._context.get('active_id', False))
            #
            # #       Remove header
            # data.pop(0)
            # today = date.today()
            # lst_date = str(today).split("-")
            # today = lst_date[2] + "/" + lst_date[1] + "/" + lst_date[0]
            # for row_vals in data:
            wb = openpyxl.load_workbook(

                filename=BytesIO(base64.b64decode(self.fastag_product))
            )
            ws = wb.active
            for row_vals in ws.iter_rows(min_row=2, max_row=None, min_col=None,

                                         max_col=None, values_only=True):
                product = self.env['product.product'].search(
                                 [('barcode', '=', row_vals[0])])
                if product:
                    print('rowvallllllllllllllllllllllll', row_vals[0])
                    product.sudo().update({'fastag_sold': 'yes'})
                    transfer_id = self.env['all.transfer.report'].search(
                        [('name', '=', row_vals[0])])
                    transfer_id.sudo().update({'rs_tag_sold': 'yes'})
                    # query1 = """UPDATE product_template
                    #                        SET fastag_sold = 'yes'
                    #                         WHERE id = %(name)s """
                    # self.env.cr.execute(query1, {'name': product.id
                    #                             })
                    # query2 = """UPDATE all_transfer_report
                    #                                        SET rs_tag_sold = 'no'
                    #                                         WHERE name = %(name)s """
                    # self.env.cr.execute(query2, {'name': product.barcode
                    #                             })
                else:
                    print(row_vals[0])
                picking_id = self.env['stock.picking'].search([])
                for rec in picking_id:
                    move_id = rec.move_ids_without_package
                    for move in move_id:
                        if move.fatag_bracode == product.barcode:
                            move.sudo().update({'fastag_sold': 'yes'})

    # def product_update_line(self):
    #     if self.fastag_product == False:
    #         raise UserError(
    #             _('Please Upload your excel first then imoport fastag'))
    #     else:
    #         terms = []
    #         # file_datas = base64.decodestring(self.fastag_product)
    #         # workbook = xlrd.open_workbook(file_contents=file_datas)
    #         #
    #         # #        Read worksheet
    #         # sheet = workbook.sheet_by_index(0)
    #         # data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r
    #         #         in range(sheet.nrows)]
    #         # data.pop(0)
    #         # today = date.today()
    #         # lst_date = str(today).split("-")
    #         # today = lst_date[2] + "/" + lst_date[1] + "/" + lst_date[0]
    #         # for row_vals in data:
    #         wb = openpyxl.load_workbook(
    #
    #             filename=BytesIO(base64.b64decode(self.fastag_product))
    #         )
    #         ws = wb.active
    #         for row_vals in ws.iter_rows(min_row=2, max_row=None, min_col=None,
    #
    #                                      max_col=None, values_only=True):
    #             product = self.env['product.product'].search(
    #                 [('barcode', '=', row_vals[0])])
    #             product_tem = self.env['product.template'].search(
    #                 [('barcode', '=', row_vals[0])])
    #             print('product', product)
    #             if product:
    #                 login = row_vals[5]
    #                 assigned_id = self.env['res.users'].sudo().search(
    #                     [('login', '=', int(login))])
    #
    #                 if assigned_id:
    #                     # print('assignee',assigned_id)
    #                     product.sudo().update({'fastag_sold': 'yes'})
    #                     product_tem.sudo().update(
    #                         {'warehouse_emp_no': assigned_id.rs_employee_id,
    #                          'product_assigned_to': assigned_id.id,
    #                          'fastag_sold': 'yes'})
    #                     transfer_id = self.env['all.transfer.report'].search(
    #                         [('name', '=', row_vals[0])])
    #                     transfer_id.sudo().write({'rs_tag_sold': 'yes'})


    def set_fastag_move_details(self):
        today = date.today()
        lst_date = str(today).split("-")
        today = lst_date[2] + "/" + lst_date[1] + "/" + lst_date[0]
        picking_ids = self.env['stock.picking'].browse(
            self._context.get('active_id', False))
        print('=================================name',picking_ids.name)
        for line in self.order_line_form:
            circle_id = self.env['rs.circle.name'].search(
                [('name', '=', line.circle_name)])
            if not circle_id:
                raise ValidationError(
                    _(' Please Create Proper Circle Name %s Is Not Given In The List!',
                      line.circle_name))
                circle_id = self.env['rs.circle.name'].create(
                    {'name': line.circle_name})
            self.env['stock.move'].sudo().create({
                'product_id': line.product_id.id,
                'fatag_bracode': line.fatag_bracode,
                'circle_name': circle_id.id,
                # 'date_of_dispatch':today,
                'consignment_no': line.consignment_no,
                'delivery_partner': line.delivery_partner,
                'emp_mobile_no': line.emp_mobile_no,
                'name': line.product_id.name,
                'product_uom': line.product_id.uom_id.id,
                'location_id': 1,
                'picking_id': line.picking_id.id,
                'location_dest_id': 1,
                'category_id': line.category_id.id
            })
            line.product_id.write(
                {'assigned_to_mob': line.picking_id.rs_destination_id.login,
                 'product_assigned_to': line.picking_id.rs_destination_id.id,
                 'warehouse_emp_no': line.picking_id.rs_destination_id.rs_employee_id})

            # self.env['transfer.report'].sudo().create({'name':line.fatag_bracode,
            #     'rs_date_dispatch':today,
            #     'assigned_id':line.picking_id.rs_destination_id.id,
            #     'rs_date_transfer':line.picking_id.scheduled_date,
            #     'rs_from_emp_no':self.user_id.rs_employee_id,
            #     'rs_to_emp_no':line.picking_id.rs_destination_id.rs_employee_id,
            #     })

        if not self.order_line_form:
            raise UserError(_('You can not confirm empty fastag !'))
        template = self.env.ref(
            'rn_vehical_class.inventory_trasfer_assigned_email_template')
        print("======", template.email_to, self.id)
        self.env['mail.template'].browse(template.id).send_mail(
            self.order_line_form[0].picking_id.id)

        return True


class MessageWizard(models.TransientModel):
    _name = 'message.wizard'

    message = fields.Char('Message', readonly=True, required=True)

    # @api.multi
    def action_ok(self):
        """ close wizard"""
        return {'type': 'ir.actions.act_window_close'}


class FastagLineForm(models.TransientModel):
    _name = 'fastag.line.form'

    product_line_id = fields.Many2one('order.line.product', string="Order Line")
    product_id = fields.Many2one('product.product', string="Products",
                                 store=True)
    picking_id = fields.Many2one('stock.picking', "Picking",
                                 required=True, ondelete='cascade')
    fatag_bracode = fields.Char(string="FASTAG BARCODE")
    circle_name = fields.Char(string="CIRCLE NAME")
    date_of_dispatch = fields.Char(string="DATE OF DISPATCH")
    consignment_no = fields.Char(string="CONSIGNMENT NO")
    delivery_partner = fields.Char(string="DELIVERY PARTNER")
    emp_mobile_no = fields.Char(string="EMP MOB NO")
    category_id = fields.Many2one('product.category', string="TAG CLASS")